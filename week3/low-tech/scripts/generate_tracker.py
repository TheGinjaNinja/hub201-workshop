"""
Hub201 Week 3 - Contact Tracker Generator
Produces: contact-tracker.xlsx
Run: python3 week3/low-tech/scripts/generate_tracker.py
"""

import os
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUTPUT_PATH = os.path.join(
    os.path.dirname(__file__), "..", "contact-tracker.xlsx"
)

# Colours
ORANGE_HEX = "EA580D"
HEADER_FILL = PatternFill("solid", fgColor="F4F4F5")
ORANGE_FILL = PatternFill("solid", fgColor=ORANGE_HEX)
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
LIGHT_ORANGE_FILL = PatternFill("solid", fgColor="FFF7ED")

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="0F172A")
ORANGE_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
LABEL_FONT = Font(name="Calibri", bold=True, size=11, color="0F172A")
BODY_FONT = Font(name="Calibri", size=11, color="0F172A")
MUTED_FONT = Font(name="Calibri", size=10, color="64748B")

THIN = Side(border_style="thin", color="E4E4E7")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell, width=None):
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.border = CELL_BORDER
    cell.alignment = Alignment(wrap_text=True, vertical="center")


def style_orange_header(cell):
    cell.font = ORANGE_FONT
    cell.fill = ORANGE_FILL
    cell.border = CELL_BORDER
    cell.alignment = Alignment(wrap_text=True, vertical="center")


def style_body(cell):
    cell.font = BODY_FONT
    cell.border = CELL_BORDER
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_contacts_sheet(wb):
    ws = wb.active
    ws.title = "Contacts"

    columns = [
        ("Name", 20),
        ("Role", 20),
        ("Company", 22),
        ("Why them (fit + trigger)", 32),
        ("Channel", 18),
        ("First-touch plan", 28),
        ("Status", 22),
        ("Interview date", 16),
        ("Notes", 30),
        ("Evidence rating", 18),
    ]

    # Header row
    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 28

    # Data rows
    for row in range(2, 32):
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row, column=col_idx, value="")
            style_body(cell)
        ws.row_dimensions[row].height = 20

    # Data validation: Status column (G = 7)
    status_dv = DataValidation(
        type="list",
        formula1='"Not yet contacted,Contacted,Replied,Booked,Completed,Referred"',
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True,
    )
    status_dv.sqref = "G2:G31"
    ws.add_data_validation(status_dv)

    # Data validation: Evidence rating column (J = 10)
    rating_dv = DataValidation(
        type="list",
        formula1='"Strong,Moderate,Weak,Assumption"',
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True,
    )
    rating_dv.sqref = "J2:J31"
    ws.add_data_validation(rating_dv)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Instructions row at top
    ws.insert_rows(1)
    instr = ws.cell(row=1, column=1,
                    value="Hub201 Week 3 Contact Tracker  |  "
                          "Target: 25 contacts minimum, 10 booked calls by "
                          "Week 4. Status and Evidence rating columns have "
                          "dropdown lists.")
    instr.font = MUTED_FONT
    instr.fill = LIGHT_ORANGE_FILL
    ws.merge_cells("A1:J1")
    instr.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 22

    ws.freeze_panes = "A3"


def build_icp_sheet(wb):
    ws = wb.create_sheet("ICP")

    def write_label(row, col, text, merged_end_col=None, orange=False):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = LABEL_FONT if not orange else ORANGE_FONT
        cell.fill = HEADER_FILL if not orange else ORANGE_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if merged_end_col:
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row, end_column=merged_end_col
            )
        return cell

    def write_value(row, col, text="", merged_end_col=None):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = BODY_FONT
        cell.border = CELL_BORDER
        cell.fill = WHITE_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if merged_end_col:
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row, end_column=merged_end_col
            )
        return cell

    # Title
    title = ws.cell(row=1, column=1,
                    value="Hub201 Week 3 - ICP Definition Sheet")
    title.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title.fill = ORANGE_FILL
    ws.merge_cells("A1:F1")
    title.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 30

    # ICP definition
    write_label(3, 1, "ICP definition (one sentence)", merged_end_col=2,
                orange=True)
    write_value(3, 3, "", merged_end_col=6)
    ws.row_dimensions[3].height = 40

    write_label(4, 1, "Target company size", merged_end_col=2)
    write_value(4, 3, "", merged_end_col=6)

    write_label(5, 1, "Target seniority/role", merged_end_col=2)
    write_value(5, 3, "", merged_end_col=6)

    write_label(6, 1, "Target industry or vertical", merged_end_col=2)
    write_value(6, 3, "", merged_end_col=6)

    write_label(7, 1, "Key trigger event", merged_end_col=2)
    write_value(7, 3, "", merged_end_col=6)

    # Credibility checklist
    ws.row_dimensions[9].height = 22
    write_label(9, 1, "Credibility checklist", merged_end_col=3,
                orange=True)
    write_label(9, 4, "Status", merged_end_col=6, orange=True)

    cred_items = [
        "Website: live, professional, explains the problem you solve",
        "LinkedIn: updated, shows relevant background",
        "Social proof: warm intro, advisor name, or reference available",
    ]
    status_options = ["Not done", "In progress", "Done"]
    for i, item in enumerate(cred_items):
        row = 10 + i
        write_label(row, 1, item, merged_end_col=3)
        cell = write_value(row, 4, "Not done", merged_end_col=6)
        ws.row_dimensions[row].height = 22

    # Credibility status dropdown
    cred_dv = DataValidation(
        type="list",
        formula1='"Not done,In progress,Done"',
        allow_blank=True,
    )
    cred_dv.sqref = "D10:F12"
    ws.add_data_validation(cred_dv)

    # Channel mix
    ws.row_dimensions[14].height = 22
    write_label(14, 1, "Channel mix target", merged_end_col=3, orange=True)
    write_label(14, 4, "% of outreach", merged_end_col=6, orange=True)

    channels = ["LinkedIn", "Events", "Warm intros", "Industry bodies"]
    for i, ch in enumerate(channels):
        row = 15 + i
        write_label(row, 1, ch, merged_end_col=3)
        cell = write_value(row, 4, "", merged_end_col=6)
        ws.row_dimensions[row].height = 20

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16


def build_affinity_sheet(wb):
    ws = wb.create_sheet("Affinity mapping")

    # Title
    title = ws.cell(row=1, column=1,
                    value="Hub201 Week 3 - Affinity Mapping")
    title.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title.fill = ORANGE_FILL
    ws.merge_cells("A1:E1")
    title.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 30

    instr = ws.cell(
        row=2, column=1,
        value="After each interview, add a row. "
              "Group rows by Theme to find patterns. "
              "One quote or observation per row."
    )
    instr.font = MUTED_FONT
    instr.fill = LIGHT_ORANGE_FILL
    ws.merge_cells("A2:E2")
    instr.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 22

    columns = [
        ("Theme", 24),
        ("Interviewee", 22),
        ("Quote or observation", 45),
        ("Hypothesis link", 28),
        ("Signal type", 22),
    ]

    for col_idx, (col_name, col_width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        style_header(cell)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[3].height = 24

    for row in range(4, 34):
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row, column=col_idx, value="")
            style_body(cell)
        ws.row_dimensions[row].height = 22

    # Signal type dropdown
    signal_dv = DataValidation(
        type="list",
        formula1='"Recency and frequency,Concrete behaviour,'
                 'Time and money,Problem-area match,Awareness level,Other"',
        allow_blank=True,
    )
    signal_dv.sqref = "E4:E33"
    ws.add_data_validation(signal_dv)

    ws.freeze_panes = "A4"


def main():
    wb = openpyxl.Workbook()
    build_contacts_sheet(wb)
    build_icp_sheet(wb)
    build_affinity_sheet(wb)

    out = os.path.abspath(OUTPUT_PATH)
    wb.save(out)
    print(f"Saved: {out}  ({os.path.getsize(out):,} bytes)")


if __name__ == "__main__":
    main()
